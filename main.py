from app import app
from flask import jsonify, request, render_template, session, make_response
from openpyxl import load_workbook
import os

		
@app.route('/pie_chart',methods = ['POST', 'GET'])
def google_pie_chart():
        if request.method == 'POST':
                colors = ['#3366cc', '#dc3912', '#ff9900', '#109618', '#990099']
                sector = ['Energy', 'Industry', 'Transport', 'Agriculture', 'Other']


                arr2 = [['Year','']]
                arr1 = []


                sectorDict = {'Energy' : 2, 'Industry' : 3, 'Transport' : 4, 'Agriculture'  : 5, 'Other' : 6}
                pollutantDict = {'PM10': 'B', 'PM2.5': 'C', 'VOC': 'D', 'CO': 'E', 'NOX': 'F', 'SO2': 'G'}
                
                year = request.form['years']

                book2 = load_workbook("final2.xlsx")
                sheet = book2[str(year)]
                

                
                pollutant = request.form['pollutant']
                
##                data = {'Task' : 'Hours per Day', 'Energy' : 11, 'Industry' : 2, 'Transport' : 2, 'Agriculture' : 2, 'Dairy & poultry' : 7,
##                'Waste water' : 7, 'others' : 7}
##                book = load_workbook("DataSheet.xlsx")
                data2 = {'Task' : 'Hours per Day'}
                

##                for col in range(97,97+sheet.max_column):  # to find pollutant name
##                    for row in "1":     # to find pollutant name
##                        cell_name = "{}{}".format(chr(col), row)        #retrived names of pollutant
##                        if name == sheet[cell_name].value:         #got the pollutant we want
##                                for row in range(2,sheet.max_row+1):    
##                                     for column in chr(col):            #Here you can add or reduce the columns
##                                             cell_name = "{}{}".format(column, row)
##                                             cell_name2 = "{}{}".format("A", row)
##                                             #print(cell_name2 + "" + cell_name)
##                                             data2[sheet[cell_name2].value] = sheet[cell_name].value



                for row in range(2,sheet.max_row+1):    
                     cell_name = "{}{}".format(pollutantDict[pollutant], row)
                     cell_name2 = "{}{}".format("A", row)
                     data2[sheet[cell_name2].value] = round(sheet[cell_name].value, 2)
                print(data2)

                return render_template('pie-chart2.html', data=data2,year=year,Pollutant=pollutant, sheet=sheet, color=colors,sector=sector )

@app.route('/services', methods=['GET', 'POST'])
def service():
        pollutantList = ['PM10', 'PM2.5', 'VOC', 'CO', 'NOX', 'SO2']
        years = [1990,1995,2000,2005,2010,2015,2020]

        Pollutant = request.args.get('type')
        resp = render_template('service-piechart.html', Pollutant=Pollutant, pollutantList=pollutantList, years=years)
        return resp

@app.route('/index')
def index():
        resp = render_template('index2.html')
        return resp

@app.route('/hello')
def hello():
   return render_template('hello2.html')


@app.route('/service2')
def service2():
        sector = ['Energy', 'Industry', 'Transport', 'Agriculture', 'Other']
        pollutant = ['PM10', 'PM2.5', 'VOC', 'CO', 'NOX', 'SO2']

        
        
        return render_template('service-trend.html', sector=sector, pollutant=pollutant)




@app.route('/timeSeries',methods = ['POST', 'GET'])
def timeSeries():
        
        arr1 = []
        arr3 = []
                
##        
##        data  = {'Task' : 'Hours per Day'}
##        data2 = [['Year',  'Expenses'],
##          ['2004',        400],
##          ['2005',        460],
##          ['2006',         1120],
##          ['2007',        540]
##        ]
##        
        sector = request.form['sector']
        pollutant = request.form['pollutant']


        sectorDict = {'Energy' : 2, 'Industry' : 3, 'Transport' : 4, 'Agriculture'  : 5, 'Other' : 6}
        pollutantDict = {'PM10': 'B', 'PM2.5': 'C', 'VOC': 'D', 'CO': 'E', 'NOX': 'F', 'SO2': 'G'}
        
        years = [1990,1995,2000,2005,2010,2015,2020]
        years2 = ['1990','1995','2000','2005','2010','2015','2020']
        
        book = load_workbook("final2.xlsx")

        arr2 = [['Year',pollutant]]

        for year in years:
                arr1.clear()
                sheet = book[str(year)]
                cell_name = "{}{}".format(pollutantDict[pollutant], sectorDict[sector])
                #data[year] = sheet[cell_name].value
                arr1.append(str(year))
                arr1.append(sheet[cell_name].value)
                arr2.append(arr1.copy())
                arr3.append(arr1.copy())
                
        return render_template('timeSeries2.html', data=arr2,data3=arr3, pollutant=pollutant, sector=sector,year=years2 )


@app.route('/first')
def first():
        return render_template('first.html')

@app.route('/trends_comb')
def trends_comb():
        prod = request.args.get('prod')
        #prod = request.form['prod']
        path = 'images/excel/'
        path2 = 'images/excel/trendbig/'
        file2 = 'images/excel/trendbig/'


        
        if prod == 'trends':
                path = path + "trends/"
                path2 = path2 + "Slide1.JPG"
                file2 = file2 + "Slide1.JPG"
        elif prod == 'piecharts':
                path = path + "pie2/"
                path2 = path2 + "Slide8.JPG"
                file2 = file2 + "Slide1.JPG"

        
        dir_list = os.listdir('static/'+path)
        print("path: " + path)
        print(dir_list)

        
        return render_template('trends_comb.html', imgs=dir_list,path=path,path2=path2,file2=file2)
   
        


@app.route('/ground_based')
def ground_based():
        return render_template('groundBased.html')

@app.route('/aeronet')
def aeronet():
        return render_template('Aeronet.html')

@app.route('/aeronetData',methods = ['POST', 'GET'])
def aeronetData():
        path = 'images/'
        data = request.form['data']
        if data == 'seasonalSizeVar':
                path = path + "aeronet/seasonal_size/"
        elif data == 'AOD_AE':
                path = path + "aeronet/aod_ae/"
        elif data == 'Annual':
                path = path + "aeronet/annual/"
                
        dir_list = os.listdir('static/'+path)
        
        
        return render_template('aeronetData.html', imgs=dir_list,path=path)


@app.route('/satellite_service')
def satellite_service():
        return render_template('satellite_service.html')

@app.route('/satellite_product/',methods = ['POST', 'GET'])
def satellite_product():

        #prod = request.args.get('prod')
        prod = request.form['prod']
        path = 'images/'

        print("prod " + prod)

        if prod == 'cam':
                path = path + 'CAMS/'
                
        if prod == 'geo':
                path = path + 'giovanni(2017-2021)2/'

        
        
        pollutant = request.form['pollutant']
        
        P_type = request.form['type']
        
        if P_type == 'Seasonal':
                path = path + "Seasonal/"+pollutant+"/"
        elif P_type == 'Average':
                path = path + "Aveg/"+pollutant+"/"
        elif P_type == 'Annual':
                path = path + "ANNUAL/"+pollutant+"/"

        
        dir_list = os.listdir('static/'+path)
        print(path)
        print(dir_list)

        
        return render_template('satellite_product.html', imgs=dir_list,path=path)
      

@app.route('/satellite_product_geo',methods = ['POST', 'GET'])
def satellite_product_geo():
        
        return render_template('satellite_product_geo.html', prod='geo')


@app.route('/satellite_product_cam',methods = ['POST', 'GET'])
def satellite_product_cam():
        
        return render_template('satellite_product_cam.html', prod='cam')

@app.route('/aboutPunjab')
def aboutPunjab():
        
        return render_template('aboutPunjab.html')
        
if __name__ == "__main__":
        app.secret_key = "123"

        app.run(debug = True, host = '0.0.0.0')
